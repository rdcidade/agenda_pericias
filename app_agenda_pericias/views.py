from .utils import enviar_pauta_peritos
from pyexpat.errors import messages
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import tbl_pericias_agendadas, tbl_unidades, tbl_usuarios, tbl_peritos
from .forms import agendar_pericias_form, cadastro_usuario_form, login_form, editar_pericia_form
from django.utils import timezone
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from .context_processors import variaveis_globais
from openpyxl import load_workbook
import logging, json
from django.db.utils import IntegrityError
from django.http import JsonResponse, HttpResponse, FileResponse
from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth
from django.db import transaction
from django.conf import settings
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from django.core.mail import send_mail
from django.utils.timezone import now
from django.urls import reverse
from django.utils.http import urlencode
from django.contrib.auth.hashers import check_password
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.hashers import make_password

logger = logging.getLogger(__name__)

# logar e autentica na aplicação
def login_form(request):
    if request.method == 'POST':
        cpf = request.POST.get('cpf')
        senha_digitada = request.POST.get('senha')

        try:
            usuario = tbl_usuarios.objects.get(cpf=cpf)
        except usuario.DoesNotExist:
            return render(request, 'login.html', {'erro': 'CPF não encontrado.'})

        # Verifica se a senha digitada bate com a senha criptografada
        if check_password(senha_digitada, usuario.senha):
            request.session['usuario_id'] = usuario.id
            request.session['usuario_nome'] = usuario.nome
            return redirect('home')  # ajuste para sua rota
        else:
            return render(request, 'agenda_pericias/erro_login.html', {'erro': 'Senha incorreta.'})

    foruns = tbl_unidades.objects.all().order_by('subsecao')
    return render(request, 'agenda_pericias/index.html',{'foruns':foruns})

def login(request):
    if request.method == 'POST':
        cpf_login = request.POST.get('cpf_login')
        senha_digitada = request.POST.get('senha_login')
        unidade = request.POST.get('unidade_login')

        try:
            usuario = tbl_usuarios.objects.get(cpf=cpf_login)
        except tbl_usuarios.DoesNotExist:
            return render(request, 'login.html', {'erro': 'CPF não encontrado.'})

        # Verifica se a senha digitada bate com a senha criptografada
        if check_password(senha_digitada, usuario.senha):
            # Armazena dados do usuário na sessão
            request.session['usuario_id'] = usuario.id
            request.session['unidade'] = unidade

            # Define as permissões dos menus conforme o perfil do usuário
            menus = {
                'disable_pericias': usuario.perito,  # True se for perito
                'disable_informacoes': usuario.perito,                
                'disable_atualizar': usuario.perito,
                'disable_area_perito': not usuario.perito,  # True se não for perito
            }
            # Redireciona para home com o contexto (sem quebrar fluxo atual)
            return render(request, 'agenda_pericias/home.html')

        else:
            return render(request, 'agenda_pericias/erro_login.html', {'erro': 'Senha incorreta.'})

    # Se for GET, exibe página inicial com dropdowns de fóruns
    foruns = tbl_unidades.objects.all().order_by('subsecao')
    return render(request, 'agenda_pericias/index.html', {'foruns': foruns})

def verificar_perito(request):
    """Verifica se o CPF digitado pertence a um perito"""
    cpf = request.GET.get('cpf_login')
    if not cpf:
        return JsonResponse({'erro': 'CPF não informado'}, status=400)
    try:
        usuario = tbl_usuarios.objects.get(cpf=cpf)
        return JsonResponse({'is_perito': bool(usuario.perito)})
    except tbl_usuarios.DoesNotExist:
        return JsonResponse({'is_perito': False})    

def user_logout(request):
    logout(request)
    messages.success(request, "Você foi desconectado com sucesso.")
    return render(request, 'agenda_pericias/index.html')

# Cadastro de usuário
def index(request):
    foruns = tbl_unidades.objects.all().order_by('subsecao')

    if request.method == 'POST':
        form = cadastro_usuario_form(request.POST, request.FILES)

        if form.is_valid():
            cpf_ = form.cleaned_data['cpf']
            subsecao_ = form.cleaned_data['unidade']

            # 🔍 Verifica duplicidade
            if tbl_usuarios.objects.filter(cpf=cpf_, unidade=subsecao_).exists():
                messages.error(request, 'Usuário já cadastrado.')
                return render(request, 'agenda_pericias/index.html', {'foruns': foruns})

            # 🔍 Busca na tabela de peritos
            perito_data = tbl_peritos.objects.filter(cpf=cpf_).first()

            # Cria o objeto usuário sem salvar ainda
            usuario = form.save(commit=False)

            # Define campos adicionais
            usuario.usuario = usuario.nome.split()[0] if usuario.nome else cpf_

            if perito_data:
                usuario.perito = True
                usuario.nome = perito_data.nome_perito
            else:
                usuario.perito = False

            # Garante que a senha será criptografada
            usuario.senha = make_password(usuario.senha)

            # 🔥 Tenta salvar
            try:
                usuario.save()
                messages.success(request, 'Usuário cadastrado com sucesso.')
            except Exception as e:
                messages.error(request, f'Erro ao salvar usuário: {e}')
                print(f"[ERRO SAVE] {e}")  # Log no terminal

            return render(request, 'agenda_pericias/index.html', {'foruns': foruns})  # Recarrega a página limpa

        else:
            # 🔍 Mostra erros de validação do formulário
            messages.error(request, f'Erro no formulário: {form.errors}')
            print(f"[FORM ERROR] {form.errors}")

            return render(request, 'agenda_pericias/index.html', {'foruns': foruns})

    # GET
    return render(request, 'agenda_pericias/index.html', {'foruns': foruns})

@csrf_exempt
def verificar_cpf(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        cpf = data.get('cpf')

        # Verifica se já existe o usuário
        if tbl_usuarios.objects.filter(cpf=cpf).exists():
            return JsonResponse({'existe_usuario': True})

        # Verifica se o CPF consta em tbl_peritos
        perito = tbl_peritos.objects.filter(cpf=cpf).first()
        if perito:
            return JsonResponse({
                'existe_usuario': False,
                'encontrado_perito': True,
                'nome_perito': perito.nome_perito
            })

        # Caso não encontre em nenhuma tabela
        return JsonResponse({
            'existe_usuario': False,
            'encontrado_perito': False
        })        

# Página inicial
def home(request):
    # Redireciona para a página de listagem de pericias filtradas
    return render(request, 'agenda_pericias/home.html')

# Página de informações
def confirmacao(request):
    return render(request, 'agenda_pericias/confirmacao.html')

# Agendar de pericias
def agendar_pericias(request):
    if request.method == "POST":
        form = agendar_pericias_form(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perícia agendada com sucesso.')
        return render(request, "agenda_pericias/agendar_pericias.html")
    else:
        peritos = tbl_peritos.objects.all().order_by('nome_perito')
        return render(request, "agenda_pericias/agendar_pericias.html", {'peritos': peritos})

# Página com a relação de pericias agendadas
def listagem_pericias(request):
    context_globals = variaveis_globais(request)
    jurisdicao = context_globals['subsecao']
    dia_atual = datetime.now().strftime('%d/%m/%Y')
    pericias = tbl_pericias_agendadas.objects.filter(situacao_pericia__icontains='Designada', jurisdicao__icontains=jurisdicao, data_pericia__icontains=dia_atual).exclude(especialidade__icontains='Social').order_by('data_marcada','hora_marcada')

    busca = request.GET.get('perito')
    busca1 = request.GET.get('processo')
    if busca:
        pericias = pericias.filter(perito__icontains=busca)
    if busca1:
        pericias = pericias.filter(processo__icontains=busca1)
    if busca and busca1:
        pericias = pericias.filter(perito__icontains=busca, processo__icontains=busca1)
    # Redireciona para a página de listagem de pericias filtradas
    return render(request, 'agenda_pericias/listagem_pericias.html', {'pericias': pericias})

def listagem_completa_pericias(request):
    context_globals = variaveis_globais(request)
    jurisdicao = context_globals['subsecao']
    pericias = tbl_pericias_agendadas.objects.filter(situacao_pericia__icontains='Designada', jurisdicao__icontains=jurisdicao).exclude(especialidade__icontains='Social').order_by('data_marcada', 'hora_marcada')

    busca = request.GET.get('search')
    if busca:
        pericias = pericias.filter(data_pericia__icontains=busca)
    # Redireciona para a página de listagem de pericias filtradas
    return render(request, 'agenda_pericias/listagem_completa_pericias.html', {'pericias': pericias})

def editar_pericia(request, id):
    pericias = tbl_pericias_agendadas.objects.get(id=id)
    if request.method == 'POST':
        form = editar_pericia_form(request.POST, instance=pericias)

        hora_entrada = request.POST.get('hora_entrada')
        hora_saida = request.POST.get('hora_saida')

        if form.is_valid():
            if hora_entrada and hora_saida:
                # Converte as horas para o formato datetime e torna as duas horas timezone-aware
                hora_entrada = timezone.make_aware(datetime.strptime(hora_entrada, "%H:%M"), timezone.get_current_timezone())
                hora_saida = timezone.make_aware(datetime.strptime(hora_saida, "%H:%M"), timezone.get_current_timezone())

                # Verifica se hora_entrada < hora_saida < timezone.now()
                if hora_entrada >= hora_saida:
                    messages.error(request, "A hora de entrada deve ser menor que a hora de saída.")
                    return render(request, 'agenda_pericias/editar_pericia.html', {'form': form, 'pericia': pericias})

                # if hora_saida >= timezone.now():
                #     messages.error(request, "A hora de saída não pode ser maior ou igual ao momento atual.")
                #     return render(request, 'agenda_pericias/editar_pericia.html', {'form': form, 'pericia': pericias})

                # Se as verificações passarem, salva os dados
                pericias.hora_entrada = hora_entrada
                pericias.hora_saida = hora_saida
                pericias.save()
                
            elif hora_entrada or hora_saida:
                # Caso um dos campos esteja preenchido, mas não o outro
                messages.error(request, "Os campos hora de entrada e hora de saída devem ser preenchidos juntos.")
                return render(request, 'agenda_pericias/editar_pericia.html', {'form': form, 'pericia': pericias})
            
            if not pericias.compareceu_pericia and pericias.hora_entrada and pericias.hora_saida:
                pericias.compareceu_pericia = True
                pericias.save()
                
            form.save()
            return redirect('listagem_pericias')

    form = editar_pericia_form(instance=pericias)
    pericia = {
        'pericia': tbl_pericias_agendadas.objects.get(id=id)
    }
    pericia = tbl_pericias_agendadas.objects.get(id=id)

    context = {
        'pericia': pericia,
        'form': form,
    }

    return render(request, 'agenda_pericias/editar_pericia.html', context)

def editar_listagem_pericia(request, id):
    pericias = tbl_pericias_agendadas.objects.get(id=id)
    if request.method == 'POST':
        form = editar_pericia_form(request.POST, instance=pericias)

        hora_entrada = request.POST.get('hora_entrada')
        hora_saida = request.POST.get('hora_saida')

        search_param = request.GET.get('search', '')

        if form.is_valid():
            if hora_entrada and hora_saida:
                # Converte as horas para o formato datetime e torna as duas horas timezone-aware
                hora_entrada = timezone.make_aware(datetime.strptime(hora_entrada, "%H:%M"), timezone.get_current_timezone())
                hora_saida = timezone.make_aware(datetime.strptime(hora_saida, "%H:%M"), timezone.get_current_timezone())

                # Verifica se hora_entrada < hora_saida < timezone.now()
                if hora_entrada >= hora_saida:
                    messages.error(request, "A hora de entrada deve ser menor que a hora de saída.")
                    return render(request, 'agenda_pericias/editar_listagem_pericia.html', {'form': form, 'pericia': pericias})

                # if hora_saida >= timezone.now():
                #     messages.error(request, "A hora de saída não pode ser maior ou igual ao momento atual.")
                #     return render(request, 'agenda_pericias/editar_pericia.html', {'form': form, 'pericia': pericias})

                # Se as verificações passarem, salva os dados
                pericias.hora_entrada = hora_entrada
                pericias.hora_saida = hora_saida
                pericias.save()
                
            elif hora_entrada or hora_saida:
                # Caso um dos campos esteja preenchido, mas não o outro
                messages.error(request, "Os campos hora de entrada e hora de saída devem ser preenchidos juntos.")
                return render(request, 'agenda_pericias/editar_listagem_pericia.html', {'form': form, 'pericia': pericias})
            
            if not pericias.compareceu_pericia and pericias.hora_entrada and pericias.hora_saida:
                pericias.compareceu_pericia = True
                pericias.save()
                
            form.save()
            query_params = {'search': search_param} if search_param else {}
            return redirect(f"{reverse('listagem_completa_pericias')}?{urlencode(query_params)}")

    form = editar_pericia_form(instance=pericias)
    pericia = {
        'pericia': tbl_pericias_agendadas.objects.get(id=id)
    }
    pericia = tbl_pericias_agendadas.objects.get(id=id)

    context = {
        'pericia': pericia,
        'form': form,
    }

    return render(request, 'agenda_pericias/editar_listagem_pericia.html', context)

def listagem_pericias_por_perito(request):
    context_globals = variaveis_globais(request)
    usuario = context_globals['usuario']
    nome = context_globals['nome']
    jurisdicao = context_globals['subsecao']
    dia_atual = datetime.now().strftime('%d/%m/%Y')
    pericias = tbl_pericias_agendadas.objects.filter(situacao_pericia__icontains='Designada', jurisdicao__icontains=jurisdicao, perito__icontains=nome, data_pericia__icontains=dia_atual).exclude(especialidade__icontains='Social')

    busca = request.GET.get('search')
    if busca:
        pericias = tbl_pericias_agendadas.objects.filter(
            data_pericia__icontains=busca, 
            situacao_pericia__icontains='designada', 
            jurisdicao__icontains=jurisdicao, 
            perito__icontains=nome).exclude(especialidade__icontains='Social')

    context = {
        'view_origem': 'dia',
        'pericias': pericias
    }

    return render(request, 'agenda_pericias/listagem_pericias_por_perito.html', context)

def listagem_pericias_por_perito_mes(request):
    context_globals = variaveis_globais(request)
    usuario = context_globals['usuario']
    nome = context_globals['nome']
    jurisdicao = context_globals['subsecao']
    mes_atual = datetime.now().strftime('%m/%Y')
    pericias = tbl_pericias_agendadas.objects.filter(
        situacao_pericia__icontains='Designada', 
        jurisdicao__icontains=jurisdicao, 
        perito__icontains=nome, 
        data_pericia__icontains=mes_atual).exclude(especialidade__icontains='Social')

    busca = request.GET.get('search')
    if busca:
        pericias = tbl_pericias_agendadas.objects.filter(
            data_pericia__icontains=busca, 
            situacao_pericia__icontains='designada', 
            jurisdicao__icontains=jurisdicao, 
            perito__icontains=nome).exclude(especialidade__icontains='Social')

    context = {
        'view_origem': 'mes',
        'pericias': pericias
    }

    return render(request, 'agenda_pericias/listagem_pericias_por_perito.html', context)

def atualizar_saida_agendamento_jef(request, id):
    try:
        agendamento = get_object_or_404(tbl_agendamento, id=id)
        if not agendamento.compareceu_agendamento:
            agendamento.compareceu_agendamento = True
        else:
            agendamento.compareceu_agendamento = True

        agendamento.save()
        return redirect('agendamento_jef')

    except tbl_agendamento.DoesNotExist:
        return JsonResponse({'error': 'O agendamento especificado não foi encontrado.'}, status=404)

def atualizar_saida_agendamento_atermacao(request, id):
    try:
        agendamento = get_object_or_404(tbl_agendamento, id=id)
        if not agendamento.compareceu_agendamento:
            agendamento.compareceu_agendamento = True
        else:
            agendamento.compareceu_agendamento = True

        agendamento.save()
        return redirect('agendamento_atermacao')

    except tbl_agendamento.DoesNotExist:
        return JsonResponse({'error': 'O agendamento especificado não foi encontrado.'}, status=404)
    
def atualizar_saida_pericia(request, id):
    pericias = get_object_or_404(tbl_pericias_agendadas, id=id)
    context_globals = variaveis_globais(request)
    usuario = context_globals['usuario']
    agora = timezone.now()
    if pericias.compareceu_pericia and pericias.hora_entrada:
        if pericias.hora_saida:
            return redirect('editar_pericia', id=id)
        else:
            pericias.hora_saida = agora
            pericias.save()
            return redirect('listagem_pericias')
    else:
        return redirect('editar_pericia', id=id)
    
def atualizar_comparecimento(request):
    agora = timezone.now()
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            pericia_id = data.get("id")
            compareceu = bool(data.get("compareceu"))

            pericia = tbl_pericias_agendadas.objects.get(id=pericia_id)

            if compareceu:
                pericia.hora_entrada = agora
                pericia.compareceu_pericia = compareceu
            else:
                if pericia.hora_entrada and pericia.hora_saida:
                    pericia.hora_entrada = None
                    pericia.hora_saida = None
                    pericia.compareceu_pericia = False
                
            pericia.save()

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método não permitido"})

# Página de Listagem Completa de Perícias
def atualizar_comparecimento_1(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            pericia_id = data.get("id")
            compareceu = bool(data.get("compareceu"))

            pericia = tbl_pericias_agendadas.objects.get(id=pericia_id)

            if pericia.hora_entrada and pericia.hora_saida:
                    pericia.hora_entrada = None
                    pericia.hora_saida = None
                    pericia.compareceu_pericia = False
                
            pericia.save()

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método não permitido"})

# Importa os registros do arquivo Excel extraído do PJ-e
def import_excel(request):
    if request.method == 'POST':
        logger.info("POST request received")
        if 'file' in request.FILES:
            try:
                excel_file = request.FILES['file']
                logger.info(f"File received: {excel_file.name}")
                if not excel_file.name.endswith('.xlsx'):
                    messages.error(request, 'O arquivo não é um Excel.')
                    return redirect('import_excel')
                wb = load_workbook(excel_file)
                sheet = wb.active

                # Mapeamento de colunas
                field_mapping = {
                    'data_pericia': 0,
                    'jurisdicao': 1,
                    'perito': 2,
                    'especialidade': 3,
                    'processo': 5,
                    'situacao_pericia': 8,
                    'periciado': 9
                }

                for row in sheet.iter_rows(min_row=15, values_only=True):
                    try:
                        situacao_pericia = row[field_mapping['situacao_pericia']]

                        # Só importa registros com situação 'Designada'
                        if situacao_pericia == 'Designada':
                            data_hora_str = row[field_mapping['data_pericia']]
                            data_hora_obj = datetime.strptime(data_hora_str, '%d/%m/%Y %H:%M:%S')
                            data = data_hora_obj.date()
                            hora = data_hora_obj.time()

                            # Extrair apenas os dois primeiros nomes do perito e periciado
                            def dois_primeiros_nomes(nome):
                                if nome:
                                    partes = nome.strip().split()
                                    return " ".join(partes[:2])  # mantém até 2 nomes
                                return ""

                            perito = dois_primeiros_nomes(row[field_mapping['perito']])
                            periciado = dois_primeiros_nomes(row[field_mapping['periciado']])

                            tbl_pericias_agendadas.objects.create(
                                data_marcada=data,
                                hora_marcada=hora,
                                data_pericia=row[field_mapping['data_pericia']],
                                jurisdicao=row[field_mapping['jurisdicao']],
                                perito=perito,
                                especialidade=row[field_mapping['especialidade']],
                                processo=row[field_mapping['processo']],
                                situacao_pericia=situacao_pericia,
                                periciado=periciado
                            )

                    except IntegrityError as e:
                        logger.warning(f"Duplicate entry found: {e}")

                messages.success(request, 'Arquivo importado com sucesso.')
                return redirect('import_excel')

            except Exception as e:
                logger.error(f"Error importing file: {e}")
                messages.error(request, f'Erro ao importar arquivo: {e}')
                return redirect('import_excel')
        else:
            logger.warning("No file found in request.FILES")
            messages.error(request, 'Nenhum arquivo foi encontrado.')
            return redirect('import_excel')

    return render(request, 'agenda_pericias/upload.html')

def pericias_nao_realizadas(request):
    ano_atual = now().year  # Obtém o ano atual
    # Agrupa as perícias por mês e especialidade, contando quantas foram realizadas em cada grupo
    pericias = (
        tbl_pericias_agendadas.objects
        .filter(compareceu_pericia=0)  # Filtra apenas as perícias não realizadas
        .filter(data_marcada__year=ano_atual)  # Filtra pelo ano atual
        .annotate(mes=TruncMonth('data_marcada'))  # Trunca a data para o mês
        .values('mes', 'especialidade')  # Seleciona o mês e a especialidade
        .annotate(total=Count('id'))  # Conta o número de perícias em cada grupo
        .order_by('mes', 'especialidade')  # Ordena por mês e especialidade
    )

    # Organizar os dados em uma lista de dicionários com totais
    pericias_por_mes = []
    mes_atual = None
    especialidades = []
    total_mes = 0

    for pericia in pericias:
        if mes_atual and pericia['mes'] != mes_atual:
            # Adiciona os dados do mês anterior ao resultado
            pericias_por_mes.append({
                'mes': mes_atual,
                'especialidades': especialidades,
                'total_mes': total_mes
            })
            # Reseta os acumuladores para o novo mês
            especialidades = []
            total_mes = 0

        mes_atual = pericia['mes']
        especialidades.append({
            'especialidade': pericia['especialidade'],
            'total': pericia['total']
        })
        total_mes += pericia['total']

    # Adiciona o último mês processado
    if mes_atual:
        pericias_por_mes.append({
            'mes': mes_atual,
            'especialidades': especialidades,
            'total_mes': total_mes
        })
    # Calcula o total de perícias no ano
    total_anual = tbl_pericias_agendadas.objects.filter(compareceu_pericia=0).aggregate(total=Count('id'))['total']

    context = {
        'pericias_por_mes': pericias_por_mes,
        'total_anual': total_anual,
        'view_origem': 'nao_realizadas'        
    }
    return render(request, 'agenda_pericias/pericias_por_especialidade.html', context)

def pericias_por_especialidade(request):
    ano_atual = now().year  # Obtém o ano atual
    # Agrupa as perícias por mês e especialidade, contando quantas foram realizadas em cada grupo
    pericias = (
        tbl_pericias_agendadas.objects
        .filter(compareceu_pericia=1)  # Filtra apenas as perícias realizadas
        .filter(data_marcada__year=ano_atual)  # Filtra pelo ano atual
        .annotate(mes=TruncMonth('data_marcada'))  # Trunca a data para o mês
        .values('mes', 'especialidade')  # Seleciona o mês e a especialidade
        .annotate(total=Count('id'))  # Conta o número de perícias em cada grupo
        .order_by('mes', 'especialidade')  # Ordena por mês e especialidade
    )

    # Organizar os dados em uma lista de dicionários com totais
    pericias_por_mes = []
    mes_atual = None
    especialidades = []
    total_mes = 0

    for pericia in pericias:
        if mes_atual and pericia['mes'] != mes_atual:
            # Adiciona os dados do mês anterior ao resultado
            pericias_por_mes.append({
                'mes': mes_atual,
                'especialidades': especialidades,
                'total_mes': total_mes
            })
            # Reseta os acumuladores para o novo mês
            especialidades = []
            total_mes = 0

        mes_atual = pericia['mes']
        especialidades.append({
            'especialidade': pericia['especialidade'],
            'total': pericia['total']
        })
        total_mes += pericia['total']

    # Adiciona o último mês processado
    if mes_atual:
        pericias_por_mes.append({
            'mes': mes_atual,
            'especialidades': especialidades,
            'total_mes': total_mes
        })
    # Calcula o total de perícias no ano
    total_anual = tbl_pericias_agendadas.objects.filter(compareceu_pericia=1).aggregate(total=Count('id'))['total']

    context = {
        'pericias_por_mes': pericias_por_mes,
        'total_anual': total_anual,
        'view_origem': 'realizadas'
    }
    return render(request, 'agenda_pericias/pericias_por_especialidade.html', context)

def pericias_por_perito(request):
    context_globals = variaveis_globais(request)
    nome = context_globals['nome']
    ano_atual = now().year
    unidade = None

    # 🔍 Verifica se o usuário é perito
    # Caso tenha um campo específico (ex: usuario.is_perito ou usuario.perfil == 'Perito'), adapte aqui
    if nome:
        nome_perito = tbl_usuarios.objects.get(nome=nome)
        is_perito = bool(nome_perito.perito)

        if is_perito:
            # Filtra apenas as perícias do perito autenticado
            pericias = (
                tbl_pericias_agendadas.objects
                .filter(compareceu_pericia=1)
                .filter(data_marcada__year=ano_atual)
                .filter(perito=nome_perito.nome)  # ajusta conforme seu modelo de usuário
                .annotate(mes=TruncMonth('data_marcada'))
                .values('mes', 'perito')
                .annotate(total=Count('id'))
                .order_by('mes', 'perito')
            )
        else:
            # Usuário comum: mostra todas as perícias
            pericias = (
                tbl_pericias_agendadas.objects
                .filter(compareceu_pericia=1)
                .filter(data_marcada__year=ano_atual)
                .annotate(mes=TruncMonth('data_marcada'))
                .values('mes', 'perito')
                .annotate(total=Count('id'))
                .order_by('mes', 'perito')
            )

        # 🔹 Organização dos dados para o template
        pericias_por_mes = []
        mes_atual = None
        peritos = []
        total_mes = 0

        for pericia in pericias:
            if mes_atual and pericia['mes'] != mes_atual:
                pericias_por_mes.append({
                    'mes': mes_atual,
                    'peritos': peritos,
                    'total_mes': total_mes
                })
                peritos = []
                total_mes = 0

            mes_atual = pericia['mes']
            peritos.append({
                'perito': pericia['perito'],
                'total': pericia['total']
            })
            total_mes += pericia['total']

        if mes_atual:
            pericias_por_mes.append({
                'mes': mes_atual,
                'peritos': peritos,
                'total_mes': total_mes
            })

        total_anual = (
            tbl_pericias_agendadas.objects
            .filter(compareceu_pericia=1)
            .aggregate(total=Count('id'))['total']
        )

        context = {
            'pericias_por_mes': pericias_por_mes,
            'total_anual': total_anual,
        }

        return render(request, 'agenda_pericias/pericias_por_perito.html', context)    

def agendamentos_por_tipo(request):
    # Agrupa os agendamentos por mês e tipo, contando quantos foram realizados em cada grupo
    agendamentos = (
        tbl_agendamento.objects
        .filter(compareceu_agendamento=1)
        .annotate(mes=TruncMonth('data_agendamento'))  # Trunca a data para o mês
        .values('mes', 'tipo_agendamento')  # Seleciona o mês e o tipo de agendamento
        .annotate(total=Count('id'))  # Conta o número de agendamentos em cada grupo
        .order_by('mes', 'tipo_agendamento')  # Ordena por mês e tipo de agendamento
    )

    # Organizar os dados em uma lista de dicionários com totais
    agendamentos_por_mes = []
    mes_atual = None
    tipos = []
    total_mes = 0

    for agendamento in agendamentos:
        if mes_atual and agendamento['mes'] != mes_atual:
            # Adiciona os dados do mês anterior ao resultado
            agendamentos_por_mes.append({
                'mes': mes_atual,
                'tipos': tipos,
                'total_mes': total_mes
            })
            # Reseta os acumuladores para o novo mês
            tipos = []
            total_mes = 0

        mes_atual = agendamento['mes']
        tipos.append({
            'tipo_agendamento': agendamento['tipo_agendamento'],
            'total': agendamento['total']
        })
        total_mes += agendamento['total']

    # Adiciona o último mês processado
    if mes_atual:
        agendamentos_por_mes.append({
            'mes': mes_atual,
            'tipos': tipos,
            'total_mes': total_mes
        })

    # Calcula o total anual
    total_anual = tbl_agendamento.objects.aggregate(total=Count('id'))['total']

    context = {
        'agendamentos_por_mes': agendamentos_por_mes,
        'total_anual': total_anual,
    }

    return render(request, 'agenda_pericias/agendamentos_por_tipo.html', context)

def listar_peritos(request):
    # Calcula a data limite (4 meses ou 120 dias atrás)
    data_limite = datetime.now() - timedelta(days=120)

    # Lista de todos os peritos nos últimos 4 meses/120 dias
    peritos = tbl_pericias_agendadas.objects.filter(data_marcada__gte=data_limite).values_list('perito', flat=True).distinct().order_by('perito')

    # Lista de todas as especialidades nos últimos 4 meses/120 dias
    lista_especialidades = tbl_pericias_agendadas.objects.filter(data_marcada__gte=data_limite).values_list('especialidade', flat=True).distinct().order_by('especialidade')

    # Peritos agrupados por especialidade nos últimos 4 meses/120 dias
    especialidades = tbl_pericias_agendadas.objects.filter(data_marcada__gte=data_limite).values('especialidade').distinct().order_by('especialidade')
    peritos_por_especialidade = {}
    for especialidade in especialidades:
        peritos_por_especialidade[especialidade['especialidade']] = tbl_pericias_agendadas.objects.filter(
            especialidade=especialidade['especialidade'], data_marcada__gte=data_limite
        ).values_list('perito', flat=True).order_by('perito').distinct()

    context = {
        'peritos': peritos,
        'peritos_por_especialidade': peritos_por_especialidade,
        'lista_especialidades': lista_especialidades,
    }

    return render(request, 'agenda_pericias/listagem_peritos.html', context)

@transaction.atomic
def atualizar_registros_pericias(request):
    # Busca todos os registros que precisam ser atualizados
    registros = tbl_pericias_agendadas.objects.all()
    
    for registro in registros:
        # Pega o valor do campo datapericia
        data_hora_str = registro.data_pericia
        
        try:
            # Converte a string para um objeto datetime
            data_hora_obj = datetime.strptime(data_hora_str, '%d/%m/%Y %H:%M')
            
            # Separa a data e a hora
            data_marcada = data_hora_obj.date()  # yyyy-mm-dd
            hora_marcada = data_hora_obj.time()  # HH:MM
            
            # Atualiza os campos do registro
            registro.data_marcada = data_marcada
            registro.hora_marcada = hora_marcada
            
            # Salva o registro atualizado
            registro.save()
            
        except ValueError as e:
            return HttpResponse(f"Erro ao processar o registro com datapericia '{data_hora_str}': {e}", status=400)
    
    return HttpResponse("Atualização dos registros concluída.")

def gerar_pdf(request):
    # Obtendo o mês filtrado do input
    mes_filtro = request.GET.get('mes', None)

    # Filtrando registros pelo mês e ano
    if mes_filtro:
        mes_ano = datetime.strptime(mes_filtro, '%Y-%m')
        registros = tbl_pericias_agendadas.objects.filter(
            data_marcada__year=mes_ano.year,
            data_marcada__month=mes_ano.month
        ).order_by('data_marcada', 'perito')
    else:
        registros = tbl_pericias_agendadas.objects.all().order_by('data_marcada', 'perito')

    # Preparando a resposta HTTP para PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_pericias.pdf"'

    # Criando o documento PDF
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    # Configurando estilos
    styles = getSampleStyleSheet()

    # Agrupando registros por data_marcada e perito
    data_perito_grupos = {}
    for reg in registros:
        chave = (reg.data_marcada, reg.perito)
        if chave not in data_perito_grupos:
            data_perito_grupos[chave] = []
        data_perito_grupos[chave].append(reg)

    # Escrevendo os registros no PDF
    for (data_marcada, perito), registros in data_perito_grupos.items():
        elements.append(Paragraph(f"Data Marcada: {data_marcada.strftime('%d/%m/%Y')} - Perito: {perito}", styles['Heading3']))
        
        # Definindo a estrutura da tabela
        table_data = [
            ['Data Perícia', 'Periciado', 'Processo', 'Especialidade']
        ]
        for reg in registros:
            table_data.append([
                reg.data_pericia,
                reg.periciado,
                reg.processo,
                reg.especialidade
            ])
        
        # Criando e estilizando a tabela com estilo zebrado
        table = Table(table_data)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Cabeçalho em cinza
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Texto do cabeçalho em branco
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordas em preto
        ])

        # Aplicando o estilo zebrado dinamicamente
        for i in range(1, len(table_data)):
            bg_color = colors.whitesmoke if i % 2 == 0 else colors.lightgrey
            table_style.add('BACKGROUND', (0, i), (-1, i), bg_color)

        table.setStyle(table_style)

        elements.append(table)
        elements.append(Paragraph("<br/>", styles['Normal']))  # Adiciona um espaço entre as tabelas

    # Construindo o PDF
    doc.build(elements)

    return response

def gerar_pdf_perito(request):
    context_globals = variaveis_globais(request)
    usuario = context_globals['usuario']

    # Filtrando registros pelo mês e ano
    mes_ano = datetime.now().month
    ano = datetime.now().year
    registros = tbl_pericias_agendadas.objects.filter(
        data_marcada__year=ano,
        data_marcada__month=mes_ano,
        perito__icontains=usuario
    ).order_by('data_marcada', 'perito')


    # Preparando a resposta HTTP para PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_pericias.pdf"'

    # Criando o documento PDF
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    # Configurando estilos
    styles = getSampleStyleSheet()

    # Agrupando registros por data_marcada e perito
    data_perito_grupos = {}
    for reg in registros:
        chave = (reg.data_marcada, reg.perito)
        if chave not in data_perito_grupos:
            data_perito_grupos[chave] = []
        data_perito_grupos[chave].append(reg)

    # Escrevendo os registros no PDF
    for (data_marcada, perito), registros in data_perito_grupos.items():
        elements.append(Paragraph(f"Data Marcada: {data_marcada.strftime('%d/%m/%Y')} - Perito: {perito}", styles['Heading3']))
        
        # Definindo a estrutura da tabela
        table_data = [
            ['Data Perícia', 'Periciado', 'Processo', 'Especialidade']
        ]
        for reg in registros:
            table_data.append([
                reg.data_pericia,
                reg.periciado,
                reg.processo,
                reg.especialidade
            ])
        
        # Criando e estilizando a tabela com estilo zebrado
        table = Table(table_data)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Cabeçalho em cinza
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Texto do cabeçalho em branco
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordas em preto
        ])

        # Aplicando o estilo zebrado dinamicamente
        for i in range(1, len(table_data)):
            bg_color = colors.whitesmoke if i % 2 == 0 else colors.lightgrey
            table_style.add('BACKGROUND', (0, i), (-1, i), bg_color)

        table.setStyle(table_style)

        elements.append(table)
        elements.append(Paragraph("<br/>", styles['Normal']))  # Adiciona um espaço entre as tabelas

    # Construindo o PDF
    doc.build(elements)

    return response

def relatorio_pericias(request):
    context_globals = variaveis_globais(request)
    jurisdicao = context_globals['subsecao']

    mes_filtro = request.GET.get('mes', None)
    mes_atual = datetime.now().month
    ano_atual = datetime.now().year
    
    
    if mes_filtro:
        mes_ano = datetime.strptime(mes_filtro, '%Y-%m')
        registros = tbl_pericias_agendadas.objects.filter(
            jurisdicao__icontains=jurisdicao,
            data_marcada__year=mes_ano.year,
            data_marcada__month=mes_ano.month
        ).order_by('data_marcada','hora_marcada', 'perito').exclude(especialidade='Assistente Social')
    else:
        mes_ano = datetime.now()
        registros = tbl_pericias_agendadas.objects.filter(
            jurisdicao__icontains=jurisdicao,
            data_marcada__year=mes_ano.year,
            data_marcada__month=mes_ano.month
            ).order_by('data_marcada','hora_marcada', 'perito').exclude(especialidade='Assistente Social')
    
        

    # Agrupando os dados por data e perito
    dados_agrupados = []
    data_perito_grupos = {}
    for reg in registros:
        chave = (reg.data_marcada, reg.perito)
        if chave not in data_perito_grupos:
            data_perito_grupos[chave] = []
        data_perito_grupos[chave].append(reg)

    for chave, grupo in data_perito_grupos.items():
        dados_agrupados.append((chave[0], chave[1], grupo))

    # Adicionando o filtro de mês ao contexto
    context = {
        'dados_agrupados': dados_agrupados,
        'mes_filtro': mes_filtro,  # Incluindo o filtro no contexto
    }

    return render(request, 'agenda_pericias/relatorio_pericias.html', context)

def enviar_pauta_view(request):
    if request.method == "POST":
        resumo = enviar_pauta_peritos()
        enviados = len(resumo['enviados'])
        sem_email = len(resumo['sem_email'])
        sem_pericias = len(resumo['sem_pericias'])
        erros = len(resumo['erros'])

        messages.success(request, f"Pautas enviadas com sucesso: {enviados}.")
        if sem_email:
            messages.warning(request, f"Peritos sem e-mail: {sem_email}.")
        if sem_pericias:
            messages.info(request, f"Peritos sem perícias no mês: {sem_pericias}.")
        if erros:
            messages.error(request, f"Ocorreram {erros} erros durante o envio.")

        return render(request, 'agenda_pericias/resumo_envio.html', {'resumo': resumo})

    return render(request, 'agenda_pericias/enviar_pauta.html')
    